from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from datetime import datetime, date


# ============================================================
# MATNNI TOZALASH
# ============================================================

def clean_text(value):

    if value is None:
        return ""

    return str(value).strip()


# ============================================================
# TUG'ILGAN SANANI FORMATLASH
# ============================================================

def format_birth_date(value):

    if value is None or value == "":
        return ""

    # Excel datetime
    if isinstance(value, datetime):

        return value.strftime(
            "%d.%m.%Y"
        )

    # Excel date
    if isinstance(value, date):

        return value.strftime(
            "%d.%m.%Y"
        )

    # Excel serial number
    if isinstance(value, (int, float)):

        try:

            converted = from_excel(
                value
            )

            if isinstance(
                converted,
                datetime
            ):

                return converted.strftime(
                    "%d.%m.%Y"
                )

            if isinstance(
                converted,
                date
            ):

                return converted.strftime(
                    "%d.%m.%Y"
                )

        except Exception:
            pass

    # Matn
    text = str(value).strip()

    formats = (
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%Y-%m-%d",
        "%Y/%m/%d",
    )

    for fmt in formats:

        try:

            converted = datetime.strptime(
                text,
                fmt
            )

            return converted.strftime(
                "%d.%m.%Y"
            )

        except ValueError:

            continue

    return text


# ============================================================
# EXCELNI O'QISH
# ============================================================

def read_students_excel(file_path):

    workbook = load_workbook(
        file_path,
        data_only=True
    )

    sheet = workbook.active

    students = []

    # --------------------------------------------------------
    # USTUNLARNI ANIQLASH
    # --------------------------------------------------------

    headers = []

    for cell in sheet[1]:

        headers.append(
            clean_text(
                cell.value
            )
        )

    print("================================")
    print("EXCEL USTUNLARI:")
    print(headers)
    print("================================")

    # --------------------------------------------------------
    # HEADER NOMLARI ORQALI USTUNNI TOPISH
    # --------------------------------------------------------

    full_name_column = None
    birth_date_column = None
    class_column = None

    for index, header in enumerate(
        headers
    ):

        header_lower = header.lower()

        # F.I.SH
        if (
            "f.i.sh" in header_lower
            or "fio" in header_lower
            or "ism" in header_lower
            or "familiya" in header_lower
        ):

            full_name_column = index

        # Tug'ilgan sana
        elif (
            "tug" in header_lower
            and "sana" in header_lower
        ):

            birth_date_column = index

        # Sinf
        elif "sinf" in header_lower:

            class_column = index

    # --------------------------------------------------------
    # SENING EXCELING UCHUN
    # --------------------------------------------------------
    #
    # A = F.I.SH
    # E = Tug'ilgan sana
    # F = Sinf
    #
    # Agar header orqali topilmasa,
    # aynan shu ustunlardan foydalanamiz.
    # --------------------------------------------------------

    if full_name_column is None:

        full_name_column = 0

    if birth_date_column is None:

        birth_date_column = 4

    if class_column is None:

        class_column = 5

    print(
        f"F.I.SH ustuni: "
        f"{full_name_column + 1}"
    )

    print(
        f"Tug'ilgan sana ustuni: "
        f"{birth_date_column + 1}"
    )

    print(
        f"Sinf ustuni: "
        f"{class_column + 1}"
    )

    print("================================")

    # --------------------------------------------------------
    # O'QUVCHILAR
    # --------------------------------------------------------

    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        if not any(
            value is not None
            for value in row
        ):

            continue

        # ----------------------------------------------------
        # F.I.SH
        # ----------------------------------------------------

        full_name = ""

        if len(row) > full_name_column:

            full_name = clean_text(
                row[full_name_column]
            )

        if not full_name:

            continue

        # ----------------------------------------------------
        # TUG'ILGAN SANA
        # ----------------------------------------------------

        birth_date = ""

        if len(row) > birth_date_column:

            birth_date = format_birth_date(
                row[birth_date_column]
            )

        # ----------------------------------------------------
        # SINF
        # ----------------------------------------------------

        class_name = ""

        if len(row) > class_column:

            class_name = clean_text(
                row[class_column]
            )

        if not class_name:

            class_name = "9-E"

        # ----------------------------------------------------
        # SAQLASH
        # ----------------------------------------------------

        student = {

            "full_name": full_name,

            "class_name": class_name,

            "birth_date": birth_date

        }

        students.append(
            student
        )

        print(
            f"{row_number}: "
            f"{full_name} | "
            f"{birth_date} | "
            f"{class_name}"
        )

    print("================================")

    print(
        f"JAMI O'QUVCHILAR: "
        f"{len(students)}"
    )

    print("================================")

    return students